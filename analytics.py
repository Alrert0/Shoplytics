# analytics.py
import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# Настройки подключения к БД (берем из config.py)
from config import DB_USER, DB_PASSWORD, DB_HOST, DB_PORT, DB_NAME

# Подключение
engine = create_engine(f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# Папки
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)


# ======================
# 1. ВИЗУАЛИЗАЦИИ (6 графиков)
# ======================

# Pie chart – распределение заказов по штатам
query_pie = """
SELECT c.customer_state, COUNT(o.order_id) as total_orders
FROM orders o
JOIN customers c ON o.customer_id = c.customer_id
GROUP BY c.customer_state;
"""
df_pie = pd.read_sql(query_pie, engine)
df_pie.set_index("customer_state").plot.pie(
    y="total_orders", autopct='%1.1f%%', legend=False, figsize=(8,8))
plt.title("Распределение заказов по штатам")
plt.ylabel("")
plt.savefig("charts/pie_chart.png")
plt.close()


# Bar chart – средняя стоимость доставки по штатам
query_bar = """
SELECT c.customer_state, AVG(oi.freight_value) as avg_freight
FROM order_items oi
JOIN orders o ON oi.order_id = o.order_id
JOIN customers c ON o.customer_id = c.customer_id
GROUP BY c.customer_state
ORDER BY avg_freight DESC
LIMIT 10;
"""
df_bar = pd.read_sql(query_bar, engine)
df_bar.plot.bar(x="customer_state", y="avg_freight", legend=False)
plt.title("Средняя стоимость доставки по штатам (Top-10)")
plt.xlabel("Штат")
plt.ylabel("Средняя доставка")
plt.savefig("charts/bar_chart.png")
plt.close()


# Horizontal bar – топ-10 категорий по выручке
query_hbar = """
SELECT p.product_category_name, SUM(oi.price) as revenue
FROM order_items oi
JOIN products p ON oi.product_id = p.product_id
GROUP BY p.product_category_name
ORDER BY revenue DESC
LIMIT 10;
"""
df_hbar = pd.read_sql(query_hbar, engine)
df_hbar.plot.barh(x="product_category_name", y="revenue", legend=False)
plt.title("Топ-10 категорий по выручке")
plt.xlabel("Выручка")
plt.ylabel("Категория товара")
plt.savefig("charts/hbar_chart.png")
plt.close()


# Line chart – количество заказов по месяцам
query_line = """
SELECT DATE_TRUNC('month', order_purchase_timestamp) as month, COUNT(*) as orders
FROM orders
GROUP BY month
ORDER BY month;
"""
df_line = pd.read_sql(query_line, engine)
df_line.plot.line(x="month", y="orders")
plt.title("Количество заказов по месяцам")
plt.xlabel("Месяц")
plt.ylabel("Количество заказов")
plt.savefig("charts/line_chart.png")
plt.close()


# Histogram – распределение цен товаров
query_hist = "SELECT price FROM order_items;"
df_hist = pd.read_sql(query_hist, engine)
df_hist['price'].plot.hist(bins=50)
plt.title("Распределение стоимости товаров")
plt.xlabel("Цена")
plt.ylabel("Частота")
plt.savefig("charts/histogram.png")
plt.close()


# Scatter plot – цена vs стоимость доставки
query_scatter = "SELECT price, freight_value FROM order_items;"
df_scatter = pd.read_sql(query_scatter, engine)
df_scatter.plot.scatter(x="price", y="freight_value", alpha=0.3)
plt.title("Цена товара vs Стоимость доставки")
plt.xlabel("Цена товара")
plt.ylabel("Стоимость доставки")
plt.savefig("charts/scatter_plot.png")
plt.close()

print("✅ Построено 6 графиков, сохранены в /charts/")


query_time = """
SELECT DATE_TRUNC('month', CAST(order_purchase_timestamp AS TIMESTAMP)) as month,
       COUNT(*) as orders
FROM orders
GROUP BY month
ORDER BY month;
"""
df_time = pd.read_sql(query_time, engine)

df_time['month'] = pd.to_datetime(df_time['month'])



fig = go.Figure()

fig.add_trace(go.Scatter(
    x=df_time['month'],
    y=df_time['orders'],
    mode='lines+markers',
    name='Заказы'
))

fig.update_layout(
    title="Заказы по месяцам",
    xaxis=dict(
        rangeselector=dict(
            buttons=list([
                dict(count=3, label="3m", step="month", stepmode="backward"),
                dict(count=6, label="6m", step="month", stepmode="backward"),
                dict(count=1, label="1y", step="year", stepmode="backward"),
                dict(step="all")
            ])
        ),
        rangeslider=dict(visible=True),
        type="date"
    ),
    yaxis_title="Количество заказов"
)

fig.show()

# ======================
# 3. ЭКСПОРТ В EXCEL
# ======================

def export_to_excel(dataframes_dict, filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # Условное форматирование числовых столбцов
        for col in range(2, ws.max_column+1):
            col_letter = chr(64+col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                ColorScaleRule(start_type="min", start_color="FFAA0000",
                               mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                               end_type="max", end_color="FF00AA00")
            )

    wb.save(filename)


dfs = {
    "Orders by State": df_pie,
    "Top Freight States": df_bar,
    "Top Categories": df_hbar,
    "Orders by Month": df_line,
    "Price Distribution": df_hist,
    "Price vs Freight": df_scatter
}


export_to_excel(dfs, "exports/report.xlsx")
print(f"✅ Создан файл report.xlsx, {len(dfs)} листа, {sum(len(df) for df in dfs.values())} строк")
